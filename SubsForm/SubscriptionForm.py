import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from datetime import datetime, date
import requests
import configparser
import os
import sys
from openpyxl import load_workbook

class SubscriptionFormApp:
    def __init__(self, master):
        self.master = master
        self.search_var = tk.StringVar()
        master.title("Formulário de Assinatura")

        if getattr(sys, 'frozen', False):
            self.script_dir = os.path.dirname(sys.executable)
        else:
            self.script_dir = os.path.dirname(os.path.abspath(__file__))

        config_file_path = os.path.join(self.script_dir, "config.ini")

        if not os.path.exists(config_file_path):
            self.create_config_file(config_file_path)

        self.config = configparser.ConfigParser()
        self.config.read(config_file_path)

        self.host = self.config.get("Form", "host")
        self.port = self.config.getint("Form", "port")

        self.form_frame = ttk.Frame(master, padding="20")
        self.form_frame.grid(row=0, column=0, sticky="nsew")
        self.form_frame.columnconfigure(1, weight=1)

        ttk.Label(self.form_frame, text="Nome do Cliente:").grid(row=0, column=0, sticky="w")
        self.client_name_entry = ttk.Entry(self.form_frame)
        self.client_name_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10))

        ttk.Label(self.form_frame, text="Data de Término (AAAA-MM-DD):").grid(row=1, column=0, sticky="w")
        self.end_date_entry = ttk.Entry(self.form_frame)
        self.end_date_entry.grid(row=1, column=1, sticky="ew", padx=(0, 10))

        ttk.Label(self.form_frame, text="Chave de Licença (opcional):").grid(row=2, column=0, sticky="w")
        self.license_key_entry = ttk.Entry(self.form_frame)
        self.license_key_entry.grid(row=2, column=1, sticky="ew", padx=(0, 10))

        # Add search entry and button for product search
        ttk.Label(self.form_frame, text="Procurar Produto:").grid(row=3, column=0, sticky="w")
        self.product_search_entry = ttk.Entry(self.form_frame, textvariable=self.search_var)
        self.product_search_entry.grid(row=3, column=1, sticky="ew", padx=(0, 10))

        self.product_search_button = ttk.Button(self.form_frame, text="Procurar", command=self.filter_products)
        self.product_search_button.grid(row=3, column=2, sticky="ew", padx=(0, 10))

        ttk.Label(self.form_frame, text="Produtos:").grid(row=4, column=0, sticky="w")

        self.product_listbox = tk.Listbox(self.form_frame, selectmode=tk.SINGLE)
        self.product_listbox.grid(row=5, column=0, columnspan=3, padx=(0, 10), pady=5, sticky="nsew")
        self.product_listbox.bind("<<ListboxSelect>>", self.on_product_select)
        self.update_product_list()

        self.product_listbox.selection_clear(0, tk.END)

        ttk.Label(self.form_frame, text="Novo Produto:").grid(row=6, column=0, sticky="w")
        self.new_product_entry = ttk.Entry(self.form_frame)
        self.new_product_entry.grid(row=6, column=1, sticky="ew", padx=(0, 10))

        self.add_product_button = ttk.Button(self.form_frame, text="Adicionar Produto", command=self.add_product)
        self.add_product_button.grid(row=6, column=2, sticky="ew", padx=(0, 10))

        self.delete_product_button = ttk.Button(self.form_frame, text="Excluir Produto", command=self.delete_product)
        self.delete_product_button.grid(row=6, column=3, sticky="ew", padx=(0, 10))

        self.buttons_frame = ttk.Frame(master)
        self.buttons_frame.grid(row=1, column=0, pady=(10, 0))
        self.buttons_frame.columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.add_button = ttk.Button(self.buttons_frame, text="Adicionar Assinatura", command=self.add_subscription)
        self.add_button.grid(row=0, column=0, pady=5, padx=(0, 5), sticky="ew")

        self.view_button = ttk.Button(self.buttons_frame, text="Visualizar Assinaturas", command=self.view_subscriptions)
        self.view_button.grid(row=0, column=1, pady=5, padx=(0, 5), sticky="ew")

        self.delete_button = ttk.Button(self.buttons_frame, text="Excluir Assinatura", command=self.delete_subscription)
        self.delete_button.grid(row=0, column=2, pady=5, padx=(0, 5), sticky="ew")

        self.renew_button = ttk.Button(self.buttons_frame, text="Renovar Assinatura", command=self.renew_subscription)
        self.renew_button.grid(row=0, column=3, pady=5, padx=(0, 5), sticky="ew")

        self.import_button = ttk.Button(self.buttons_frame, text="Importar do Excel", command=self.import_from_excel)
        self.import_button.grid(row=0, column=4, pady=5, padx=(0, 5), sticky="ew")

        self.check_api_status()

    def check_api_status(self):
        try:
            api_url = f"http://{self.host}:{self.port}/is_api_online"
            response = requests.get(api_url)
            response.raise_for_status()
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
            self.disable_buttons()
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao conectar à API: {e}")
            self.disable_buttons()

    def disable_buttons(self):
        self.add_button.config(state=tk.DISABLED)
        self.view_button.config(state=tk.DISABLED)
        self.delete_button.config(state=tk.DISABLED)
        self.renew_button.config(state=tk.DISABLED)
        self.add_product_button.config(state=tk.DISABLED)
        self.delete_product_button.config(state=tk.DISABLED)
        self.import_button.config(state=tk.DISABLED)

    def add_subscription(self):
        client_name = self.client_name_entry.get().strip()
        selected_product = self.product_listbox.get(tk.ACTIVE)
        end_date_str = self.end_date_entry.get().strip()
        license_key = self.license_key_entry.get().strip() or None

        if not all([client_name, selected_product, end_date_str]):
            self.handle_error("Erro", "Preencha todos os campos, exceto Chave de Licença.")
            return

        try:
            self.check_api_status()

            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            if end_date < datetime.today().date():
                raise ValueError("A data de término não pode estar no passado.")

            api_url = f"http://{self.host}:{self.port}/add_subscription"
            new_subscription = {
                "client_name": client_name,
                "product_name": selected_product,
                "end_date": end_date_str,
                "license_key": license_key
            }
            response = requests.post(api_url, json=new_subscription)
            response.raise_for_status()
            messagebox.showinfo("Sucesso", "Assinatura adicionada com sucesso.")
        except ValueError as e:
            self.handle_error("Erro", str(e))
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao adicionar assinatura: {e}")

    
    def view_subscriptions(self):
        try:
            self.check_api_status()

            api_url = f"http://{self.host}:{self.port}/view_subscriptions"
            response = requests.get(api_url)
            response.raise_for_status()

            subscriptions = response.json()

            if subscriptions:
                # Attach the original index to each subscription
                for i, sub in enumerate(subscriptions):
                    sub['original_index'] = i + 1

                self.view_window = tk.Toplevel(self.master)
                self.view_window.title("Visualizar Assinaturas")

                screen_width = self.view_window.winfo_screenwidth()
                screen_height = self.view_window.winfo_screenheight()
                x = (screen_width - self.view_window.winfo_reqwidth()) // 2
                y = (screen_height - self.view_window.winfo_reqheight()) // 2
                self.view_window.geometry("+{}+{}".format(x, y))

                search_frame = ttk.Frame(self.view_window)
                search_frame.grid(row=0, column=0, columnspan=5, padx=10, pady=5, sticky="ew")

                search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
                search_entry.grid(row=0, column=0, padx=(0, 5), sticky="ew")

                search_button = ttk.Button(search_frame, text="Procurar", command=self.filter_subscriptions)
                search_button.grid(row=0, column=1, padx=(5, 0))

                sort_button = ttk.Button(search_frame, text="Ordenar Alfabeticamente", command=self.sort_subscriptions)
                sort_button.grid(row=0, column=2, padx=(5, 0))

                restore_button = ttk.Button(search_frame, text="Restaurar Ordem Original", command=self.restore_subscriptions)
                restore_button.grid(row=0, column=3, padx=(5, 0))

                view_frame = ttk.Frame(self.view_window, padding="10")
                view_frame.grid(row=1, column=0, columnspan=5, sticky="nsew")

                view_frame.columnconfigure(0, weight=1)
                view_frame.rowconfigure(0, weight=1)

                tree_columns = ("Client Name", "Product Name", "End Date", "License Key", "Original Index")
                self.tree = ttk.Treeview(view_frame, columns=tree_columns, show="headings")

                for col in tree_columns:
                    self.tree.heading(col, text=col)
                    self.tree.column(col, anchor="center")

                vsb = ttk.Scrollbar(view_frame, orient="vertical", command=self.tree.yview)
                self.tree.configure(yscroll=vsb.set)
                self.tree.grid(row=0, column=0, sticky="nsew")
                vsb.grid(row=0, column=1, sticky="ns")

                self.filtered_subscriptions = subscriptions
                for sub in subscriptions:
                    values = (sub["client_name"], sub["product_name"], sub["end_date"], sub["license_key"], sub['original_index'])
                    self.tree.insert("", "end", values=values)

            else:
                messagebox.showinfo("Sem Assinaturas", "Nenhuma assinatura encontrada.")

        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao buscar assinaturas: {e}")

    def filter_subscriptions(self):
        search_term = self.search_var.get().strip().lower()
        self.tree.delete(*self.tree.get_children())
        filtered_subscriptions = [
            sub for sub in self.filtered_subscriptions
            if search_term in sub["client_name"].lower() or search_term in sub["product_name"].lower()
        ]
        for sub in filtered_subscriptions:
            values = (sub["client_name"], sub["product_name"], sub["end_date"], sub["license_key"], sub['original_index'])
            self.tree.insert("", "end", values=values)

    def sort_subscriptions(self):
        sorted_subscriptions = sorted(self.filtered_subscriptions, key=lambda x: x["client_name"].lower())
        self.tree.delete(*self.tree.get_children())
        for sub in sorted_subscriptions:
            values = (sub["client_name"], sub["product_name"], sub["end_date"], sub["license_key"], sub['original_index'])
            self.tree.insert("", "end", values=values)

    def restore_subscriptions(self):
        self.tree.delete(*self.tree.get_children())
        for sub in self.filtered_subscriptions:
            values = (sub["client_name"], sub["product_name"], sub["end_date"], sub["license_key"], sub['original_index'])
            self.tree.insert("", "end", values=values)

    def delete_subscription(self):
        client_name = self.client_name_entry.get().strip()
        selected_product = self.product_listbox.get(tk.ACTIVE)

        if not all([client_name, selected_product]):
            self.handle_error("Erro", "Preencha o nome do cliente e selecione um produto.")
            return

        try:
            self.check_api_status()

            api_url = f"http://{self.host}:{self.port}/delete_subscription"
            subscription = {"client_name": client_name, "product_name": selected_product}
            response = requests.post(api_url, json=subscription)
            response.raise_for_status()
            messagebox.showinfo("Sucesso", "Assinatura excluída com sucesso.")
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao excluir assinatura: {e}")

    def renew_subscription(self):
        client_name = self.client_name_entry.get().strip()
        selected_product = self.product_listbox.get(tk.ACTIVE)
        additional_days = simpledialog.askinteger("Renovar Assinatura", "Número de dias para adicionar:")

        if not all([client_name, selected_product, additional_days]):
            self.handle_error("Erro", "Preencha o nome do cliente, selecione um produto e forneça os dias adicionais.")
            return

        try:
            self.check_api_status()

            api_url = f"http://{self.host}:{self.port}/renew_subscription"
            subscription = {
                "client_name": client_name,
                "product_name": selected_product,
                "additional_days": additional_days
            }
            response = requests.post(api_url, json=subscription)
            response.raise_for_status()
            messagebox.showinfo("Sucesso", "Assinatura renovada com sucesso.")
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao renovar assinatura: {e}")

    def import_from_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])

        if not file_path:
            return

        try:
            self.check_api_status()

            workbook = load_workbook(filename=file_path, data_only=True)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                client_name, product_name, end_date_str, license_key = row

                if not client_name or not product_name or not end_date_str:
                    continue

                end_date = date.fromisoformat(end_date_str)

                api_url = f"http://{self.host}:{self.port}/add_subscription"
                new_subscription = {
                    "client_name": client_name,
                    "product_name": product_name,
                    "end_date": end_date.isoformat(),
                    "license_key": license_key
                }
                response = requests.post(api_url, json=new_subscription)
                response.raise_for_status()

            messagebox.showinfo("Sucesso", "Assinaturas importadas com sucesso do Excel.")
        except ValueError as e:
            self.handle_error("Erro", f"Erro ao processar o arquivo Excel: {e}")
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao importar assinaturas: {e}")

    def handle_error(self, title, message):
        messagebox.showerror(title, message)

    def add_product(self):
        new_product = self.new_product_entry.get().strip()

        if not new_product:
            self.handle_error("Erro", "Nome do produto não pode estar vazio.")
            return

        try:
            self.check_api_status()

            api_url = f"http://{self.host}:{self.port}/add_product"
            product_data = {"product_name": new_product}
            response = requests.post(api_url, json=product_data)
            response.raise_for_status()

            messagebox.showinfo("Sucesso", "Produto adicionado com sucesso.")
            self.update_product_list()
            self.new_product_entry.delete(0, tk.END)
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao adicionar produto: {e}")

    def delete_product(self):
        selected_product = self.product_listbox.get(tk.ACTIVE)

        if not selected_product:
            self.handle_error("Erro", "Selecione um produto para excluir.")
            return

        try:
            self.check_api_status()

            api_url = f"http://{self.host}:{self.port}/delete_product"
            product_data = {"product_name": selected_product}
            response = requests.post(api_url, json=product_data)
            response.raise_for_status()

            messagebox.showinfo("Sucesso", "Produto excluído com sucesso.")
            self.update_product_list()
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao excluir produto: {e}")

    def on_product_select(self, event):
        selected_product = self.product_listbox.get(tk.ACTIVE)
        if selected_product:
            self.new_product_entry.delete(0, tk.END)
            self.new_product_entry.insert(0, selected_product)

    def filter_products(self):
        search_term = self.search_var.get().strip().lower()
        self.product_listbox.delete(0, tk.END)
        try:
            self.check_api_status()

            api_url = f"http://{self.host}:{self.port}/get_products"
            response = requests.get(api_url)
            response.raise_for_status()

            products = response.json()

            # Filter products based on search term
            filtered_products = [product for product in products if search_term in product.lower()]

            for product in filtered_products:
                self.product_listbox.insert(tk.END, product)

        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao buscar produtos: {e}")

    def update_product_list(self):
        self.search_var.set("")  # Reset the search field
        self.product_listbox.delete(0, tk.END)
        try:
            self.check_api_status()

            api_url = f"http://{self.host}:{self.port}/get_products"
            response = requests.get(api_url)
            response.raise_for_status()

            products = response.json()

            for product in products:
                self.product_listbox.insert(tk.END, product)

        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao buscar produtos: {e}")

    def create_config_file(self, config_file_path):
        config = configparser.ConfigParser()

        config["Form"] = {
            "host": "localhost",
            "port": "5000"
        }

        with open(config_file_path, "w") as config_file:
            config.write(config_file)

if __name__ == "__main__":
    root = tk.Tk()
    app = SubscriptionFormApp(root)
    root.mainloop()
